#!/usr/bin/env python

import openpyxl
from openpyxl.utils import get_column_letter

from FileReader import FileReader


class TCPacketIllustrator:
    """
    Render the packet structure visually
    """
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.Workbook()
        self.max_cells_per_row = 32
        self.file_reader = FileReader(self.file_path)

    def create_sheet(self, data, sheet_index):
        """Create a new sheet and populate it with data."""
        sheet = self.workbook.create_sheet(title=f"Sheet {sheet_index + 1}")
        self.add_headers(sheet)
        self.populate_sheet(sheet, data)
        self.set_column_widths(sheet)
        self.set_column_heights(sheet)

    def add_headers(self, sheet):
        """Add header row with numbers from 0 to 31."""
        for col in range(self.max_cells_per_row):
            sheet.cell(row=1, column=col + 1, value=col)

    def set_column_widths(self, sheet):
        """Set the width of each column to approximately ..."""
        for col in range(1, self.max_cells_per_row + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 4

    def set_column_heights(self, sheet):
        """Set the height of each column to approximately ..."""
        for row in range(1, self.max_cells_per_row + 1):
            sheet.row_dimensions[row].height = 40

    def populate_sheet(self, sheet, data):
        """Populate the sheet with merged cells based on the dictionary data."""
        current_row = 2  # Start from the second row after the header
        current_col = 1

        for string, size in data.items():
            num_cells = size

            while num_cells > 0:
                cells_to_merge = min(
                    num_cells, self.max_cells_per_row - current_col + 1
                )
                sheet.merge_cells(
                    start_row=current_row,
                    start_column=current_col,
                    end_row=current_row,
                    end_column=current_col + cells_to_merge - 1,
                )
                cell = sheet.cell(row=current_row, column=current_col)
                cell.value = string

                current_col += cells_to_merge
                num_cells -= cells_to_merge

                if current_col > self.max_cells_per_row:
                    current_row += 1
                    current_col = 1

    def save_workbook(self, output_file):
        """Save the workbook to the specified output file."""
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]  # Remove the default sheet
        self.workbook.save(output_file)

    def create_excel(self, output_file):
        """Main method to create the Excel file from the text file."""
        for index, data in enumerate(self.file_reader.read_dictionary()):
            self.create_sheet(data, index)
        self.save_workbook(output_file)

# todo - add CLI
# if __name__ == "__main__":
#     excel_creator = TCPacketIllustrator("data.sample")
#     excel_creator.create_excel("sample.xlsx")
